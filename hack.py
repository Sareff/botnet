#!/usr/bin/python3
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from requests import *
import json, re, os

class Users:
  users = {}

  def loadUsers(self, wb):
    sheet = wb['Table 1']

    for i in range(2, sheet.max_row+1):
      self.users.update( {sheet.cell(row=i, column=3).value: sheet.cell(row=i, column=4).value} )

  def __len__(self):
    return len(self.users)



def learn(login, password):

  s = Session()

  response = s.get(url = "http://dist.kait20.ru/login/index.php")
  response = BeautifulSoup(response.text, 'html.parser')
  inputs = response.find('input', {'name': 'logintoken'})
  token = inputs.get('value')

  response = s.post(url = "http://dist.kait20.ru/login/index.php", data = {
          'anchor': '',
          'logintoken': token,
          'username': login,
          'password': password
      })
  pattern = re.compile(r"M.cfg = (\{.*?});", re.MULTILINE | re.DOTALL)
  response = BeautifulSoup(response.text, 'html.parser')
  script = response.find('script', text=pattern)
  obj = {}

  if script:
    obj = pattern.search(script.text).group(1)
    obj = json.loads(obj)
  else:
    print("Cannot prepare script to get SESSKEY")
    return 
  
  sesskey = obj['sesskey']

  data = [{
              "index":0,
              "methodname":"core_course_get_enrolled_courses_by_timeline_classification",
              "args":
              {
                  "offset":0,
                  "limit":24,
                  "classification":"all",
                  "sort":"fullname"
              }
          }]

  response = s.post(url = "http://dist.kait20.ru/lib/ajax/service.php", params = {
    'sesskey': sesskey,
    'info': 'core_cource_get_enrolled_courses_by_timeline_classification'
  }, json = data) 

  try:
    answer = response.json()
    URL = answer[0]['data']['courses'][0]['viewurl']
  except:
    print('Cannot find any course in dashboard\nShutdown...')
    return

  response = s.get(url = URL)
  response = BeautifulSoup(response.text, 'html.parser')
  ids = response.find_all('input', {'name':'id'})
  idValues = []

  for inputs in ids:
    idValues.append(inputs.get('value'))
  
  for idValue in idValues:
    response = s.post(url = "http://dist.kait20.ru/course/togglecompletion.php", data={
      'id':idValue,
      'completionstate':1,
      'fromajax':1,
      'sesskey':sesskey
    })


bots = Users()
i = input("Paste database you want to initalize (in the same partition): ")
bots.loadUsers(load_workbook(os.getcwd()+'/'+i))

for login, password in bots.users.items():
  print(login, password)
  learn(login, password)
