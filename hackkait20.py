from openpyxl import load_workbook
from bs4 import BeautifulSoup
from requests import *
import json
import re

def learn(login, password):
    s = Session()

    URL = "http://dist.kait20.ru/login/index.php"

    r = s.get(url = URL)

    data = r.text
    soup = BeautifulSoup(data, 'html.parser')
    inputs = soup.find("input", {"name":"logintoken"})
    token = inputs.get('value')

    data = {'anchor': '',
            'logintoken': token,
            'username': login,
            'password': password}

    r = s.post(url = URL, data = data)

    ans = BeautifulSoup(r.text, 'html.parser')
    pattern = re.compile(r"M.cfg = (\{.*?});", re.MULTILINE | re.DOTALL)
    script = ans.find("script", text=pattern)
    obj = {}

    if script:
        obj = pattern.search(script.text).group(1)
        obj = json.loads(obj)
    else: 
        print("No script")
        return 

    sesskey = obj["sesskey"]
    print(sesskey)

    URL  = "http://dist.kait20.ru/lib/ajax/service.php"

    params = {'sesskey': sesskey,
              'info': 'core_course_get_enrolled_courses_by_timeline_classification'}

    data = [{
                "index":0,
                "methodname":"core_course_get_enrolled_courses_by_timeline_classification",
                "args":
                {
                    "offset":0,
                    "limit":24,
                    "classification":"all",
                    "sort":"fullname"
                }
            }]

    r = s.post(url = URL, params = params, json = data)

    ans = r.json()
    URL = ans[0]["data"]["courses"][0]["viewurl"]

    r = s.get(url = URL)

    soup = BeautifulSoup(r.text, "html.parser")
    ids = soup.find_all("input", {"name":"id"})
    idValues = []
    for inputt in ids:
        idValues.append(inputt.get('value'))

    for idValue in idValues:
        data = {
                "id": idValue,
                "completionstate": 1,
                "fromajax": 1,
                "sesskey": sesskey
                }
        URL = "http://dist.kait20.ru/course/togglecompletion.php"
        r = s.post(url = URL, data = data)

wb = load_workbook('./1.xlsx')
sheet = wb['Table 1']
users = {}
for i in range(2, 36):
    users.update( {sheet.cell(row=i, column=3).value: sheet.cell(row=i, column=4).value} )

for login, password in users.items():
    print(login, password)
    learn(login, password)
    
    

